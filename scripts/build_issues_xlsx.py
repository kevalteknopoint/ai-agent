#!/usr/bin/env python3
"""build_issues_xlsx.py <findings.json> <output.xlsx>

Deterministic tracker builder shared by every code-scan analyzer agent.
Analyzer agents emit findings as plain JSON (cheap, no binary formatting
tokens spent); this script turns that JSON into the standard single-sheet
xlsx tracker (frozen header, autofilter, severity conditional formatting).

findings.json shape:
{
  "issues": [
    {
      "id": "001", "severity": 5, "severityLabel": "Critical",
      "file": "path/to/file.ext", "line": 24, "category": "XSS & context handling",
      "problem": "...", "impact": "...",
      "currentCode": "...", "recommendedFix": "...", "optimizedExample": "...",
      "complexity": "Low|Med|High", "estHours": 1.5
    }, ...
  ]
}
"""
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.stderr.write(
        "openpyxl is required. Install with: pip3 install --quiet openpyxl\n"
    )
    sys.exit(1)

COLUMNS = [
    ("Issue ID", "id"),
    ("Severity", "severity"),
    ("Severity Label", "severityLabel"),
    ("File", "file"),
    ("Line", "line"),
    ("Category", "category"),
    ("Problem", "problem"),
    ("Impact", "impact"),
    ("Current Code", "currentCode"),
    ("Recommended Fix", "recommendedFix"),
    ("Optimized Example", "optimizedExample"),
    ("Complexity", "complexity"),
    ("Est. Hours", "estHours"),
]

SEVERITY_FILLS = {
    5: "FFC7CE",  # red
    4: "FFD9B3",  # orange
    3: "FFF2CC",  # yellow
    2: "E2EFDA",  # light green
    1: "D9D9D9",  # gray
}


def main():
    if len(sys.argv) != 3:
        sys.stderr.write("usage: build_issues_xlsx.py <findings.json> <output.xlsx>\n")
        sys.exit(1)

    findings_path, output_path = sys.argv[1], sys.argv[2]
    with open(findings_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    issues = data.get("issues", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

    for col_idx, (label, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, issue in enumerate(issues, start=2):
        for col_idx, (_label, key) in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=issue.get(key))

    ws.freeze_panes = "A2"
    last_row = max(len(issues) + 1, 1)
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    sev_col_letter = get_column_letter(2)  # Severity is column B
    if len(issues) > 0:
        sev_range = f"{sev_col_letter}2:{sev_col_letter}{last_row}"
        for level, color in SEVERITY_FILLS.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.conditional_formatting.add(
                sev_range,
                CellIsRule(operator="equal", formula=[str(level)], fill=fill),
            )

    widths = [10, 9, 14, 40, 6, 22, 45, 40, 35, 35, 35, 10, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    print(json.dumps({"written": output_path, "issueCount": len(issues)}))


if __name__ == "__main__":
    main()
